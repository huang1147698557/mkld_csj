import openpyxl
wb = openpyxl.load_workbook('/Users/Apple/procalc5/procalc5.proflute.xlsx', data_only=True)
ws = wb.active
print('Sheet:', ws.title, 'rows:', ws.max_row, 'cols:', ws.max_column)
print('--- Header ---')
for c in range(1, ws.max_column+1):
    val = ws.cell(row=1, column=c).value
    print('  col%d (idx%d): %s' % (c, c-1, val))
print('--- Row 2 ---')
for c in range(1, min(ws.max_column+1, 35)):
    v = ws.cell(row=2, column=c).value
    if v is not None:
        print('  col%d (idx%d): %s' % (c, c-1, v))
